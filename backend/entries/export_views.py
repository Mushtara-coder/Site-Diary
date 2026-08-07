from io import BytesIO

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.views import View
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import SiteEntry, Issue
from projects.models import Project


class ExportPDFView(View):
    def get(self, request):
        org = request.user.organization
        if not org:
            return HttpResponse("No organization", status=400)

        entries = self._get_queryset(request, org)
        project_name = self._get_project_name(request)

        html_string = render_to_string("entries/diary_report.html", {
            "entries": entries,
            "project_name": project_name,
            "date_from": entries.order_by("date").first().date if entries.exists() else timezone.now().date(),
            "date_to": entries.order_by("-date").first().date if entries.exists() else timezone.now().date(),
            "generated_at": timezone.now(),
        })

        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_string.encode("utf-8")), result)
        if pdf.err:
            return HttpResponse("PDF generation error", status=500)

        response = HttpResponse(result.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="site_diary_report.pdf"'
        return response

    def _get_queryset(self, request, org):
        qs = SiteEntry.objects.filter(organization=org)

        project_id = request.query_params.get("project")
        if project_id:
            qs = qs.filter(project_id=project_id)

        date_from = request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(date__gte=date_from)

        date_to = request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(date__lte=date_to)

        return qs.select_related("project", "user").prefetch_related(
            "work_items", "deliveries", "plans", "issues"
        ).order_by("-date", "-created_at")

    def _get_project_name(self, request):
        project_id = request.query_params.get("project")
        if project_id:
            try:
                return Project.objects.get(id=project_id).name
            except Project.DoesNotExist:
                pass
        return "All Projects"


class ExportExcelView(View):
    def get(self, request):
        org = request.user.organization
        if not org:
            return HttpResponse("No organization", status=400)

        entries = self._get_queryset(request, org)
        project_name = self._get_project_name(request)

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        self._build_summary_sheet(wb, entries, project_name, request)
        self._build_entries_sheet(wb, entries, header_font, header_fill, header_align)
        self._build_work_items_sheet(wb, entries, header_font, header_fill, header_align)
        self._build_deliveries_sheet(wb, entries, header_font, header_fill, header_align)
        self._build_plans_sheet(wb, entries, header_font, header_fill, header_align)
        self._build_issues_sheet(wb, entries, header_font, header_fill, header_align)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="site_diary_report.xlsx"'
        return response

    def _get_queryset(self, request, org):
        qs = SiteEntry.objects.filter(organization=org)

        project_id = request.query_params.get("project")
        if project_id:
            qs = qs.filter(project_id=project_id)

        date_from = request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(date__gte=date_from)

        date_to = request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(date__lte=date_to)

        return qs.select_related("project", "user").prefetch_related(
            "work_items", "deliveries", "plans", "issues"
        ).order_by("-date", "-created_at")

    def _get_project_name(self, request):
        project_id = request.query_params.get("project")
        if project_id:
            try:
                return Project.objects.get(id=project_id).name
            except Project.DoesNotExist:
                pass
        return "All Projects"

    def _apply_header_style(self, ws, row_num, num_cols, header_font, header_fill, header_align):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _auto_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    def _build_summary_sheet(self, wb, entries, project_name, request):
        ws = wb.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35

        summary_data = [
            ("Site Diary Report", ""),
            ("", ""),
            ("Project", project_name),
            ("Total Entries", entries.count()),
            ("", ""),
            ("Issues - Low", Issue.objects.filter(entry__in=entries, severity="LOW").count()),
            ("Issues - Medium", Issue.objects.filter(entry__in=entries, severity="MEDIUM").count()),
            ("Issues - High", Issue.objects.filter(entry__in=entries, severity="HIGH").count()),
            ("Issues - Critical", Issue.objects.filter(entry__in=entries, severity="CRITICAL").count()),
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)

        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    def _build_entries_sheet(self, wb, entries, header_font, header_fill, header_align):
        ws = wb.create_sheet("Entries")
        headers = ["Date", "Project", "Weather", "Personnel", "Recorded By"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._apply_header_style(ws, 1, len(headers), header_font, header_fill, header_align)

        for row_idx, entry in enumerate(entries, start=2):
            ws.cell(row=row_idx, column=1, value=entry.date.isoformat())
            ws.cell(row=row_idx, column=2, value=entry.project.name)
            ws.cell(row=row_idx, column=3, value=entry.weather)
            ws.cell(row=row_idx, column=4, value=entry.personnel)
            ws.cell(row=row_idx, column=5, value=f"{entry.user.first_name} {entry.user.last_name}")

        self._auto_width(ws)

    def _build_work_items_sheet(self, wb, entries, header_font, header_fill, header_align):
        ws = wb.create_sheet("Work Items")
        headers = ["Entry Date", "Description", "Quantity", "Unit"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._apply_header_style(ws, 1, len(headers), header_font, header_fill, header_align)

        row_idx = 2
        for entry in entries:
            for item in entry.work_items.all():
                ws.cell(row=row_idx, column=1, value=entry.date.isoformat())
                ws.cell(row=row_idx, column=2, value=item.description)
                ws.cell(row=row_idx, column=3, value=float(item.quantity) if item.quantity else "")
                ws.cell(row=row_idx, column=4, value=item.unit)
                row_idx += 1

        self._auto_width(ws)

    def _build_deliveries_sheet(self, wb, entries, header_font, header_fill, header_align):
        ws = wb.create_sheet("Deliveries")
        headers = ["Entry Date", "Material", "Quantity", "Supplier", "Condition"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._apply_header_style(ws, 1, len(headers), header_font, header_fill, header_align)

        row_idx = 2
        for entry in entries:
            for d in entry.deliveries.all():
                ws.cell(row=row_idx, column=1, value=entry.date.isoformat())
                ws.cell(row=row_idx, column=2, value=d.material)
                ws.cell(row=row_idx, column=3, value=float(d.quantity) if d.quantity else "")
                ws.cell(row=row_idx, column=4, value=d.supplier)
                ws.cell(row=row_idx, column=5, value=d.condition)
                row_idx += 1

        self._auto_width(ws)

    def _build_plans_sheet(self, wb, entries, header_font, header_fill, header_align):
        ws = wb.create_sheet("Plans")
        headers = ["Entry Date", "Activity", "Expected Date"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._apply_header_style(ws, 1, len(headers), header_font, header_fill, header_align)

        row_idx = 2
        for entry in entries:
            for p in entry.plans.all():
                ws.cell(row=row_idx, column=1, value=entry.date.isoformat())
                ws.cell(row=row_idx, column=2, value=p.activity)
                ws.cell(row=row_idx, column=3, value=p.expected_date.isoformat() if p.expected_date else "")
                row_idx += 1

        self._auto_width(ws)

    def _build_issues_sheet(self, wb, entries, header_font, header_fill, header_align):
        ws = wb.create_sheet("Issues")
        headers = ["Entry Date", "Severity", "Description"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._apply_header_style(ws, 1, len(headers), header_font, header_fill, header_align)

        row_idx = 2
        for entry in entries:
            for i in entry.issues.all():
                ws.cell(row=row_idx, column=1, value=entry.date.isoformat())
                ws.cell(row=row_idx, column=2, value=i.severity)
                ws.cell(row=row_idx, column=3, value=i.description)
                row_idx += 1

        self._auto_width(ws)
